from django.shortcuts import render, redirect
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.http.request import HttpRequest
from django.urls import reverse
from .forms import TaskForm
from django.contrib import messages
import datetime
from .forms import  MaquinaForm
from .models import Task
from .models import Maquina
from tasks.models import Task
from .forms2 import  ConsumosForm
from .models import Consumos
from django.db import models
from django.views.generic import View
import weasyprint
from reportlab.pdfgen import canvas
from io import BytesIO
from xhtml2pdf import pisa
from django.template.loader import get_template 
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageTemplate, Frame
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Image
from django.db.models import Q
import tempfile
from django.utils import timezone
import csv
import weasyprint
from django.template.loader import render_to_string  
from django.core.paginator import Paginator
import tablib
from tablib import formats
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


@login_required
def taskList(request):
    
    search = request.GET.get('search')
    filter = request.GET.get('filter')
    tasksDoneRecently = Task.objects.filter(done='done', updated_at__gt=datetime.datetime.now()-datetime.timedelta(days=30)).count()
    tasksDone = Task.objects.filter(done='done', user=request.user).count()
    tasksDoing = Task.objects.filter(done='doing', user=request.user).count()

    if search:
        tasks = Task.objects.filter(title__icontains=search, user=request.user)
    elif filter:
        tasks = Task.objects.filter(done=filter, user=request.user)
    else:
        tasks_list = Task.objects.all().order_by('-created_at').filter(user=request.user)

        paginator = Paginator(tasks_list, 3)

        page = request.GET.get('page')
        tasks = paginator.get_page(page)

    return render(request, 'tasks/list.html', 
        {'tasks':tasks, 'tasksrecently': tasksDoneRecently, 'tasksdone': tasksDone, 'tasksdoing': tasksDoing })

@login_required
def taskView(request, id):
    task = get_object_or_404(Task, pk=id)
    return render(request, 'tasks/task.html', {'task': task})

@login_required
def taskList(request):
    
    if request.method == 'POST':
        form = TaskForm(request.POST)
        
        if form.is_valid():
            task = form.save(commit=False)
            task.done = 'doing'
            task.user = request.user
            task.save()
            return redirect('/')
    else:
        form = TaskForm()
        return render(request, 'tasks/list.html', {'form': form})

@login_required
def taskView(request, id):
    task = get_object_or_404(Task, pk=id)
    return render(request, 'tasks/task.html', {'task': task})

@login_required
def newTask(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        
        if form.is_valid():
            task = form.save(commit=False)
            task.done = 'doing'
            task.user = request.user
            task.save()
            return redirect('/')
    else:
        form = TaskForm()
        return render(request, 'tasks/addtask.html', {'form': form})

@login_required
def editTask(request, id):
    task = get_object_or_404(Task, pk=id)
    form = TaskForm(instance=task)

    if(request.method == 'POST'):
        form = TaskForm(request.POST, instance=task)

        if(form.is_valid()):
            task.save()
            return redirect('/')
        else:
            return render(reuquest, 'task/edittask.html', {'form': form, 'task': task})
   

@login_required
def deleteTask(request, id):
    task = get_object_or_404(Task, pk=id)
    task.delete()

    messages.info(request, 'Tarefa deletada com sucesso.')

    return redirect('/')

@login_required
def changeStatus(request, id):
    task = get_object_or_404(Task, pk=id)

    if(task.done == 'doing'):
        task.done = 'done'
    else:
        task.done = 'doing'

    task.save()

    return redirect('/')

@login_required
def helloWorld(request):
    return HttpResponse('Hello World!')
    
@login_required
def yourName(request, name):
    return render(request, 'tasks/yourname.html', {'name':name})



def maquina(request):
    filter_modelo = request.GET.get("filter_modelo")
    filter_tipo = request.GET.get("filter_tipo")
    filter_placa = request.GET.get("filter_placa")
    filter_chassi = request.GET.get("filter_chassi")
    filter_data = request.GET.get("filter_data")
    
    data = {'db': Maquina.objects.all()}

    if filter_modelo:
        data['db'] = data['db'].filter(modelo=filter_modelo)
    if filter_tipo:
        data['db'] = data['db'].filter(tipo=filter_tipo)
    if filter_placa:
        data['db'] = data['db'].filter(placa=filter_placa)
    if filter_chassi:
        data['db'] = data['db'].filter(chassi=filter_chassi)  # Corrigi a coluna
    if filter_data:
        data['db'] = data['db'].filter(data=filter_data)  # Corrigi a coluna

    return render(request, 'tasks/maquina.html', data)


def form(request):
    data = {}
    data['form'] = MaquinaForm()
    return render(request, 'tasks/form.html', data)
    
def create(request):
    form = MaquinaForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Salvo com sucesso')  # Adicione a mensagem de sucesso
        return redirect('form')

    # Se o formulário não for válido, você pode adicionar uma mensagem de erro se desejar
    # messages.error(request, 'Ocorreu um erro ao salvar o objeto.')

    return render(request, 'tasks/form.html', {'form': form})

    
def view(request, pk):
    data = {}
    data['db'] = Maquina.objects.get(pk=pk)
    return render(request, 'tasks/view.html', data)


def edit(request, pk):
    data ={}
    data['db'] = Maquina.objects.get(pk=pk)
    data['form'] = MaquinaForm(instance=data['db'])
    return render(request, 'tasks/form.html', data)

def update(request, pk):
    data = {}
    data['db'] = Maquina.objects.get(pk=pk)
    if request.method == 'POST':
      form = MaquinaForm(request.POST, instance=data['db'])
      if form.is_valid():
        form.save()
        return redirect('maquina')

      else:
          
          data['form'] = form
    else:
        
        form = MaquinaForm(instance=data['db'])

    data['form'] = form

    return render(request, 'tasks/maquina.html', data)

def delete(request, pk):
    db = Maquina.objects.get(pk=pk)
    db.delete()
    return redirect('maquina')


def consumo(request):
    filter_operador = request.GET.get("filter_operador")
    filter_data = request.GET.get("filter_data")
    filter_litragem = request.GET.get("filter_litragem")
    filter_valor = request.GET.get("filter_valor")
    filter_kminicial = request.GET.get("filter_kminicial")
    filter_kmfinal = request.GET.get("filter_kmfinal")

    data = {'db': Consumos.objects.all()}

    if filter_operador:
        data['db'] = data['db'].filter(operador=filter_operador)

    if filter_data:
        data['db'] = data['db'].filter(data=filter_data)

    if filter_litragem:
        data['db'] = data['db'].filter(litragem=filter_litragem)

    if filter_valor:
        data['db'] = data['db'].filter(valor=filter_valor)
    if filter_kminicial:
        data['db'] = data['db'].filter(valor=filter_kminicial)
    if filter_kmfinal:
        data['db'] = data['db'].filter(valor=filter_kmfinal)

    return render(request, 'tasks/consumo.html', data)

def form2(request):
    data = {}
    data['form'] = ConsumosForm
    return render(request, 'tasks/form2.html', data)


def cre(request):
    form = ConsumosForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Salvo com sucesso')  # Adicione a mensagem de sucesso
        return redirect('form2')

    # Se o formulário não for válido, você pode adicionar uma mensagem de erro se desejar
    # messages.error(request, 'Ocorreu um erro ao salvar o objeto.')

    return render(request, 'tasks/form2.html', {'form': form})

def view2(request, pk):
    data = {}
    data['db'] = Consumos.objects.get(pk=pk)
    return render(request, 'tasks/view2.html', data)

def edit2(request, pk):
    data = {}
    data['db'] = Consumos.objects.get(pk=pk)
    data['form'] = ConsumosForm(instance=data['db'])
    return render(request, 'tasks/form2.html', data)

def update2(request, pk):
    data = {}
    data['db'] = Consumos.objects.get(pk=pk)
    form = ConsumosForm(request.POST or None, instance=data['db'])
    if form.is_valid():
        form.save()
        return redirect('consumo') 
def delete2(request, pk):
    db = Consumos.objects.get(pk=pk)
    db.delete()
    return redirect('consumo')  



def pdf(request):
    filter_modelo = request.GET.get('filter_modelo')
    filter_tipo = request.GET.get('filter_tipo')
    filter_placa = request.GET.get('filter_placa')
    filter_chassi = request.GET.get('filter_chassi')
    filter_data = request.GET.get('filter_data')

    maquina = Maquina.objects.all()  # Comece com todos os objetos

    if filter_modelo:
        maquina = maquina.filter(modelo__icontains=filter_modelo)
    if filter_tipo:
        maquina = maquina.filter(tipo__icontains=filter_tipo)
    if filter_placa:
        maquina = maquina.filter(placa__icontains=filter_placa)
    if filter_chassi:
        maquina = maquina.filter(chassi__icontains=filter_chassi)
    if filter_data:
        maquina = maquina.filter(data__icontains=filter_data)

    context = {'db': maquina}

    html_index = render_to_string('tasks/pdf.html', context)

    weasyprint_html = weasyprint.HTML(string=html_index, base_url=request.build_absolute_uri('/'))
    
    pdf = weasyprint_html.write_pdf(stylesheets=[
    weasyprint.CSS(string='body { font-family: serif; }'),
    weasyprint.CSS(string='table { width: 100%; border-collapse: collapse; }'),
    weasyprint.CSS(string='table, th, td { border: 1px solid black; }'),
    weasyprint.CSS(string='body { font-family: serif} img {margin: 10px; width: 50px;}'),
    # Outras regras CSS
    ])

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Relatorio de maquina{datetime.datetime.now()}.pdf'
    response['Content-Transfer-Encoding'] = 'binary'

    with tempfile.NamedTemporaryFile(delete=True) as output:
        output.write(pdf)
        output.flush()
        output.seek(0)
        response.write(output.read())

    return response

def pdf1(request):
    filter_operador = request.GET.get('filter_operador')
    filter_data= request.GET.get('filter_data')
    filter_litragem = request.GET.get('filter_litragem')
    filter_valor = request.GET.get('filter_valor')
    filter_kminicial = request.GET.get("filter_kminicial")
    filter_kmfinal = request.GET.get("filter_kmfinal")

    consumo = Consumos.objects.all()  # Comece com todos os objetos

    if filter_operador:
        consumo = consumo.filter(operador__icontains=filter_operador)
    if filter_data:
        consumo= consumo.filter(data__icontains=filter_data)
    if filter_litragem:
        consumo = consumo.filter(litragem__icontains=filter_litragem)
    if filter_valor:
        consumo = consumo.filter(valor__icontains=filter_valor)
    if filter_kminicial:
        consumo = consumo.filter(valor__icontains=filter_kminicial)
    if filter_kmfinal:
        consumo = consumo.filter(valor__icontains=filter_kmfinal)
    
    context = {'db': consumo}

    html_index = render_to_string('tasks/pdf1.html', context)

    weasyprint_html = weasyprint.HTML(string=html_index, base_url=request.build_absolute_uri('/'))
    
    pdf = weasyprint_html.write_pdf(stylesheets=[
    weasyprint.CSS(string='body { font-family: serif; }'),
    weasyprint.CSS(string='table { width: 100%; border-collapse: collapse; }'),
    weasyprint.CSS(string='table, th, td { border: 1px solid black; }'),
    weasyprint.CSS(string='body { font-family: serif} img {margin: 10px; width: 50px;}'),
    # Outras regras CSS
    ])

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Relatorio de consumo{datetime.datetime.now()}.pdf'
    response['Content-Transfer-Encoding'] = 'binary'

    with tempfile.NamedTemporaryFile(delete=True) as output:
        output.write(pdf)
        output.flush()
        output.seek(0)
        response.write(output.read())

    return response


def cvs(request):
    response = HttpResponse(content_type='application/ms-excel')
    current_time = timezone.now()
    filename = f'Maquina_{current_time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    maquinas = Maquina.objects.all()

    filter_modelo = request.GET.get("filter_modelo")
    filter_tipo = request.GET.get("filter_tipo")
    filter_placa = request.GET.get("filter_placa")
    filter_chassi = request.GET.get("filter_chassi")
    filter_data = request.GET.get("filter_data")

    if filter_modelo:
        maquinas = maquinas.filter(modelo__icontains=filter_modelo,)
    if filter_tipo:
        maquinas = maquinas.filter(tipo__icontains=filter_tipo)
    if filter_placa:
        maquinas = maquinas.filter(placa__icontains=filter_placa)
    if filter_chassi:
        maquinas = maquinas.filter(chassi__icontains=filter_chassi)
    if filter_data:
        maquinas = maquinas.filter(data__icontains=filter_data)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define um objeto Border para todas as células
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
   

      

    # Escreva os títulos das colunas e aplique as bordas
    worksheet.append(['Modelo', 'Tipo', 'Placa', 'Chassi', 'Data'])
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border

    # Escreva os dados e aplique as bordas
    for maquina in maquinas:
        formatted_date = maquina.data.strftime('%d/%m/%Y')
        row = [maquina.modelo, maquina.tipo, maquina.placa, maquina.chassi, formatted_date]
        
        # Aplica bordas às células
        for col in worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row):
            for cell in col:
                cell.border = border

        worksheet.append(row)

    workbook.save(response)

    return response


def cvs1(request):
    response = HttpResponse(content_type='application/ms-excel')
    current_time = timezone.now()
    filename = f'Consumos_{current_time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Obtenha os dados de Consumos a serem exportados
    consumo = Consumos.objects.all()

    filter_operador = request.GET.get("filter_operador")
    filter_data = request.GET.get("filter_data")
    filter_litragem = request.GET.get("filter_litragem")
    filter_valor = request.GET.get("filter_valor")
    filter_kminicial = request.GET.get("filter_kminicial")
    filter_kmfinal = request.GET.get("filter_kmfinal")

    # Aplicar filtros aos dados de Consumos
    if filter_operador:
        consumo = consumo.filter(operador__icontains=filter_operador)
    if filter_data:
        consumo = consumo.filter(data__icontains=filter_data)
    if filter_litragem:
        consumo = consumo.filter(litragem__icontains=filter_litragem)
    if filter_valor:
        consumo = consumo.filter(valor__icontains=filter_valor)
    if filter_kminicial:
        consumo = consumo.filter(kminicial__icontains=filter_kminicial)
    if filter_kmfinal:
        consumo = consumo.filter(kmfinal__icontains=filter_kmfinal)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define um objeto Border para todas as células
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
   
    # Escreva os títulos das colunas e aplique as bordas
    worksheet.append(['Operador', 'Data', 'Litragem', 'Valor', 'Km Inicial', 'Km Final'])
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border

    # Escreva os dados e aplique as bordas
    for consumo in consumo:
        formatted_date = consumo.data.strftime('%d/%m/%Y')
        row = [consumo.operador, formatted_date, consumo.litragem, consumo.valor, consumo.kminicial, consumo.kmfinal]
        
        # Aplica bordas às células
        for col in worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row):
            for cell in col:
                cell.border = border

        worksheet.append(row)

    workbook.save(response)

    return response